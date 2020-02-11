#! python3
""" Tworzy plik xlsx + wykres """

import openpyxl
from openpyxl.styles import Font, NamedStyle, Color, PatternFill
from openpyxl.styles.colors import RED
from openpyxl.chart import BarChart, Series, Reference
import os
import shelve

fpath = os.getcwd()
filename = 'example.xlsx'
file_path = os.path.join(fpath, filename)


def open_excel_file():
    """ Sprawdza czy arkusz istnieje. Jeśli nie -- tworzy go """
    # Jeśli plik xlsx istnieje:
    if os.path.exists(file_path):
        # otwórz go
        work_book = openpyxl.load_workbook(file_path)
        work_book.active = 0
        work_sheet = work_book.active
        work_sheet['B1'] = 'Wiek'
        work_sheet.freeze_panes = 'A2'
    else:
        # w przeciwnym wypadku utwórz i nadaj arkuszowi nazwę
        work_book = openpyxl.Workbook()
        work_sheet = work_book.create_sheet('NewSheet', index=0)
        work_sheet.freeze_panes = 'A2'
    return work_book, work_sheet


def input_a_person(i_persons):
    """ Wprowadź osoby """

    person_dct = {}     # słownik dla aktualnych wpisów

    def check_person_dct(c_person):
        """ Sprawdza, czy osoba istnieje już w słowniku """
        if c_person not in i_persons.keys() and c_person not in person_dct.keys():
            age = int(input('Wiek: '))
            person_dct.setdefault(person, age)
        else:
            print('Osoba juz istnieje!')

    while (person := (input('Imię i Nazwisko: '))) != '':
        try:
            check_person_dct(person)    # Sprawdzić, czy osoba już istnieje
        except ValueError:
            print('Wartość wieku! Nie zapisano osoby!')
    return person_dct


def write_to_excel(family, w_work_sheet):
    """ Zapis słownika do arkusza kalkulacyjnego """

    for key, val in family.items():
        w_work_sheet.cell(row=w_work_sheet.max_row + 1, column=1, value=key)
        w_work_sheet.cell(row=w_work_sheet.max_row, column=2, value=val)


def read_an_excel(r_work_sheet):
    """ Wczytuje zawartość do słownika """
    r_dct = {}
    max_inputs = r_work_sheet.max_row

    if r_work_sheet.max_row > 1:
        for row in range(1, r_work_sheet.max_row + 1):
            if r_work_sheet['A' + str(row)].value is not None:
                r_dct.setdefault(r_work_sheet['A' + str(row)].value, r_work_sheet['B' + str(row)].value)
    else:
        print('Brak danych do wyświetlenia. Wprowadź dane.')
    return r_dct, max_inputs


def adjust_column_width(m, a_work_sheet):
    """Ustawia właściwą szerokość kolumny A"""

    a_work_sheet.column_dimensions['A'].width = m + 2


def check_length_of_dct_keys(inputs):
    """Sprawdza i zwraca najdłuższy klucz słownika"""
    lst = list(inputs.keys())
    m = []
    for i, val in enumerate(lst):
        m.append(len(lst[i]))
    m_max = max(m)
    return m_max


wb, ws = open_excel_file()
persons, max_number_of_inputs = read_an_excel(ws)
len1 = check_length_of_dct_keys(persons)
new_persons = input_a_person(persons)
len2 = check_length_of_dct_keys(new_persons)
cell_width = max(len1, len2)
adjust_column_width(cell_width, ws)
write_to_excel(new_persons, ws)
mx = ws.max_row

# Utworzenie wykresu
chartObj = BarChart()
chartObj.type = 'col'
chartObj.style = 10
chartObj.title = 'Wiek użytkowników'
chartObj.y_axis.title = 'Wiek'
chartObj.x_axis.title = 'Użytkownik'

data = Reference(ws, min_col=2, min_row=1, max_row=mx, max_col=2)
cats = Reference(ws, min_col=1, min_row=2, max_row=mx)
chartObj.add_data(data, titles_from_data=True)
chartObj.set_categories(cats)
chartObj.shape = 3
# cs = wb.create_chartsheet()
# cs.add_chart(chartObj)
ws.add_chart(chartObj, "C10")

# Zapisz arkusz
wb.save(file_path)

file_path = os.path.join(fpath, 'max_data')

shelfFile = shelve.open(file_path)
shelfFile['max_number_of_inputs'] = max_number_of_inputs
shelfFile.close()

shelfFile = shelve.open(file_path)
mx = list(shelfFile.values())
print(f'MX = {mx}')
shelfFile.close()
# TODO: 1. Wczytać arkusz
# TODO: 2. Zliczyć zawartość -- ilość wierszy -- i załadować do słownika?

