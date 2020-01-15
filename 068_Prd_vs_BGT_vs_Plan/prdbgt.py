#! python3

"""
1. Czyta trzy pliki Excela
2. pobiera konkretne komórki
3. wykonuje operacje
4. Zapisuje wartości
"""

import openpyxl
from datetime import datetime
import calendar
import sys
import os

# Kolejność argumentów: 1. Plan produkcji   2. Budżet   3. Produkcja vs BGT vs Plan   4. Wykonanie prod. narastająco'

if len(sys.argv) < 5:
    print('Brak wystarczających argv. Koniec programu.')
    sys.exit()

path = r'F:\Py\068_Prd_vs_BGT_vs_Plan\\'    # Ścieżka dostępu do katalogu z plikami

name_plan = os.path.join(path, sys.argv[1])     # łączy path z nazwą pliku
name_bgt = os.path.join(path, sys.argv[2])
name_prd_gra_mdf = os.path.join(path, sys.argv[3])
# name_act_prd = os.path.join(path, sys.argv[4])     # wykonanie produkcji narastająco

wb_plan = openpyxl.load_workbook(name_plan, data_only=True)             # ładuje plik Plan_produkcji.xlsx.
# data_only=True -- pobiera wartość z komórki, a nie formuły
sheet_plan = wb_plan['Plan']                              # Chcemy pracować na danych z arkusza Plan produkcji
wb_bgt = openpyxl.load_workbook(name_bgt)                      # ładuje plik Budżet.xlsx
sheet_bgt = wb_bgt['Arkusz1']                                        # Chcemy pracować na danych z arkusza Budżet
wb_prd_vs_bgt = openpyxl.load_workbook(name_prd_gra_mdf)  # ładuje plik Produkcja_vs_BGT.xlsx
sheet_prd_vs_bgt = wb_prd_vs_bgt['Arkusz1']                         # Chcemy pracować na danych z arkusza Arkusz1.
# W Arkusz1 zapiszemy wszystkie wyliczenia
# sheet_act_prd = openpyxl.load_workbook(name_act_prd)    # ładuje plik WYKONANIE PRD

day = datetime.now().day - 1        # dzień miesiąca. Posłuży do dalszych obliczeń 'powinno być / jest'
month = datetime.now().month    # Obecny miesiąc. Posłuży do wyliczeń: plan/ilość dni * day
year = datetime.now().year      # Do wykorzystania w monthrange -- ilość dni bieżącego miesiąca

number_of_days_current_month_tmp = calendar.monthrange(year, month)
number_of_days_current_month = number_of_days_current_month_tmp[1]  # Z pary Rok - Miesiąc, pobiera Miesiąc


# Funkcja zapisująca wyliczone wartości 'powinno być' budżetowe i planowe w Produkcja_vs_Budżet_vs_Plan
def write_to_excel_file(lista, col):
    # zapisanie MFC:
    sheet_prd_vs_bgt.cell(row=9, column=col, value=lista[0])
    # zapisanie pozostałych:
    x = 1  # MFC jest pod [0]
    for num in range(12, 15):
        sheet_prd_vs_bgt.cell(row=num, column=col, value=lista[x])
        x += 1


lst = []
for i in range(3, 15):      # zapis miesięcy do listy, z której później odszukam pozycję komórek dla bieżącego miesiąca
    lst.append(sheet_bgt.cell(row=4, column=i).value)

id_current_month = lst.index(month) + 3     # w tym Arkuszu jest na pozycji +3

# Słownik {'produkt': wartość budżetowa}
dane = {'mfc': 0, 'filmy': 0, 'folie': 0, 'obrzeża': 0}
dane_lst = [sheet_bgt.cell(row=13, column=id_current_month).value,  # Odczyt MFC z pliku budżetowego i zapis do listy
            sheet_bgt.cell(row=21, column=id_current_month).value,  # Odczyt FILMY z pliku budżetowego i zapis do listy
            sheet_bgt.cell(row=25, column=id_current_month).value,  # Odczyt FOLIE z pliku budżetowego i zapis do listy
            sheet_bgt.cell(row=29, column=id_current_month).value]  # Odczyt OBRZEŻA z pliku budżetowego i zapis

for i, key in enumerate(dane.keys()):
    dane[key] = dane_lst[i]     # aktualizacja słownika wartościami z excela
    i += 1

# ----------------------------------------------------------------------------
# Przeliczenie Powinno być wg BGT do dziś
tmp_lst = []
for k, v in dane.items():
    tmp_lst.append(round(dane[k] * day / number_of_days_current_month, 1))

# Zapisanie danych w Excelu - Produkcja_vs_Budżet
write_to_excel_file(tmp_lst, 5)

# Zapisanie arkusza/pliku
wb_prd_vs_bgt.save(name_prd_gra_mdf)

# ----------------------------------------------------------------------------
# Obliczenie forecast'u na podstawie planu produkcji. Dane z 'plan produkcji' Zapis w 'Produkcja_vs_BGT'
mfc = int(sheet_plan['F31'].value)
foil = int(sheet_plan['F47'].value)
film = int(sheet_plan['F56'].value)
edges = int(sheet_plan['F63'].value)

plan_dct_tmp = {'mfc': mfc, 'film': film, 'foil': foil, 'edges': edges}
plan_lst = []
for k, v in plan_dct_tmp.items():
    plan_lst.append(round(plan_dct_tmp[k] * day / number_of_days_current_month, 1))

# Zapisanie danych w Excelu - Produkcja_vs_Budżet
write_to_excel_file(plan_lst, 4)

# Zapisanie arkusza/pliku
wb_prd_vs_bgt.save(name_prd_gra_mdf)
